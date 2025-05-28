"""
    Microscopy Data Organiser

    Processes .CSV files from microscopy analysis by categorising them into
    Excel sheets based on binary variables and generating summary statistics.

    Supports multiple input files and creates unique output Excel files
    for each.


"""

#
# Setup
#

import os
import logging
import argparse
import pandas as pd
from openpyxl import load_workbook
VERSION = "2.0.0"

# Set up logging
logging.basicConfig(level=logging.WARNING)


#
# Parse command line arguments for the script
#
# Provide one or more input .CSV files
# Specify columns to summarise using the -c option
#
# Reutrns parsed command line arguments
#

def load_commandline():
    parser = argparse.ArgumentParser(description='Microscopy Data Organiser')

    # Input file argument
    helpstr = 'Input file(s) from microscopy analysis'
    parser.add_argument('input', type=str, nargs='+', help=helpstr)

    # Remove rows where Children_Nuclei_Count is 0
    helpstr = 'Remove rows where Children_Nuclei_Count is 0'
    parser.add_argument('-r', '--remove-zero-nuclei',
                        action='store_true', help=helpstr)

    # Version information
    helpstr = 'Display version and exit'
    out = f"This is\n Patrick Jane version {VERSION}"
    parser.add_argument('-V', '--version', action='version', help=helpstr,
                        version=out)

    # Select columns for summary
    helpstr = 'Comma-separated list of columns to summarise (default: AreaShape_Area, Children_Nuclei_Count)'
    parser.add_argument(
        '-c', '--columns',
        nargs='+',
        default=["AreaShape_Area", "Children_Nuclei_Count"],
        help="List of columns to summarise (separated by spaces)"
        )

    # Specify max bin number for summary histogram
    helpstr = 'Value for largest bin in histogram (default: 5)'
    parser.add_argument(
        '--histogram',
        type=int,
        default=5,
        help="Number of nuclei in the maximum bin for the histogram"
        )

    args = parser.parse_args()

    return (args)


#
# Create unique output names
#


used_sheet_names = set()


def get_unique_sheet_name(name, max_length=25):
    base_name = name[:max_length]
    unique_name = base_name
    counter = 1

    # Avoid overwriting if output with same name exists
    while unique_name in used_sheet_names:
        suffix = f"_{counter}"
        unique_name = (base_name[:max_length - len(suffix)] + suffix)
        counter += 1

    used_sheet_names.add(unique_name)
    return unique_name

#
# Defining the Jane function to make the summary pages
#


def jane(outfile, summary_columns):
    file_size = os.path.getsize(outfile)
    print("")
    print(f"File '{outfile}' exists. Size: {file_size} bytes. I know who the killer is I'm just not going to tell you for another 45-50 minutes")
    print("")

    xls = pd.ExcelFile(outfile, engine='openpyxl')

    # Summary data containers
    summary_data = {col: [] for col in summary_columns}

    # Iterate through the sheets
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name)

        for col in summary_columns:
            if col in df.columns:
                summary_data[col].append(df[col].rename(sheet_name))

    # Error checking if no data is found for specific columns
    for col, data in summary_data.items():
        if not data:
            print(f"No data found in any sheet for column '{col}'.")

    # Create summary DataFrames
    summary_dfs = {
        col: pd.concat(data, axis=1) if data else None
        for col, data in summary_data.items()
    }

    # Load existing sheets to avoid overwriting them
    with pd.ExcelWriter(outfile, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for col, summary_df in summary_dfs.items():
            if summary_df is not None:
                # Truncate the sheet name so openpxyl doesn't get angry
                sheet_name = get_unique_sheet_name(f"{col} Summary")
                summary_df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("")
    # Put the right column names in not just defaults
    print("Successfully created 'Area Summary' and 'Nuclei Summary' sheets. I'm going to sit in a derelict attic now instead of doing police work")
    print("")


#
# The Patrick function
#

# Parse command-line arguments
args = load_commandline()

# Required columns
required_columns = ['ObjectNumber', 'Classify_Mononucleated', 'Classify_Infected'] + args.columns

for input_file in args.input:  # Loop through all input files

    # Check if the file is real
    if not os.path.exists(input_file):
        logging.error(f"Input file '{input_file}' not found. Skipping.")
        continue

    # Read the CSV file, take the first row as the header
    try:
        df = pd.read_csv(input_file, header=0, usecols=required_columns)
    except pd.errors.EmptyDataError:
        logging.error(f"No columns to parse from {input_file}")
        continue
    except ValueError:
        logging.error(f"Required columns not found in {input_file} file :(")
        continue

    # Strip whitespace
    df.columns = df.columns.str.strip()

    # Remove rows where Children_Nuclei_Count is 0 (if option is specified)
    if args.remove_zero_nuclei:
        if "Children_Nuclei_Count" in df.columns:
            initial_row_count = len(df)
            df = df[df["Children_Nuclei_Count"] != 0]
            final_row_count = len(df)
            print(f"Removed {initial_row_count - final_row_count} rows where Children_Nuclei_Count was 0. They were not of use to me just like how I have no need for concrete evidence in my cases")
        else:
            print("Column 'Children_Nuclei_Count' not found. Skipping row removal.")

    # Total cells = number of rows
    total_cells = df.shape[0]

    # Total nuclei = sum of Children_Nuclei_Count
    total_nuclei = df['Children_Nuclei_Count'].sum()

    # Create summary of infected cells and nuclei by nuclear status
    infected_summary = df[df['Classify_Infected'] == 1]

    # Mononucleated
    mono_df = infected_summary[infected_summary['Classify_Mononucleated'] == 1]
    mono_cell_count = mono_df.shape[0]
    mono_nuclei_count = mono_df['Children_Nuclei_Count'].sum()

    # Multinucleated
    multi_df = infected_summary[infected_summary['Classify_Mononucleated'] == 0]
    multi_cell_count = multi_df.shape[0]
    multi_nuclei_count = multi_df['Children_Nuclei_Count'].sum()

    # Totals
    total_infected_cell_count = mono_cell_count + multi_cell_count
    total_infected_nuclei_count = mono_nuclei_count + multi_nuclei_count

    # Summary DataFrame for infected only
    infected_summary_df = pd.DataFrame({
        "Category": ["Mononucleated Infected", "Multinucleated Infected", "Total Infected"],
        "Infected Cell Count": [mono_cell_count, multi_cell_count, total_infected_cell_count],
        "Infected Nuclei Count": [mono_nuclei_count, multi_nuclei_count, total_infected_nuclei_count]
    })

    # Overall totals (uninfected included)
    total_cells = df.shape[0]
    total_nuclei = df['Children_Nuclei_Count'].sum()

    # Summary DataFrame for infected only, with percentages
    infected_summary_df = pd.DataFrame({
        "Category": ["Mononucleated Infected", "Multinucleated Infected", "Total Infected"],
        "Infected Cell Count": [mono_cell_count, multi_cell_count, total_infected_cell_count],
        "Infected Nuclei Count": [mono_nuclei_count, multi_nuclei_count, total_infected_nuclei_count]
    })

    # Calculate percentages
    infected_summary_df["% of Total Cells"] = (infected_summary_df["Infected Cell Count"] / total_cells * 100).round(2)
    infected_summary_df["% of Total Nuclei"] = (infected_summary_df["Infected Nuclei Count"] / total_nuclei * 100).round(2)

    # Add total row (no percentages)
    total_row = pd.DataFrame({
        "Category": ["Total incl uninfected"],
        "Infected Cell Count": [total_cells],
        "Infected Nuclei Count": [total_nuclei],
        "% of Total Cells": [""],
        "% of Total Nuclei": [""]
    })

    # Combine
    infected_summary_df = pd.concat([infected_summary_df, total_row], ignore_index=True)

    # Categorize the data according to infection and nuclei
    categories = {
        'Mononucleated_Infected': df[(df['Classify_Mononucleated'] == 1) & (df['Classify_Infected'] == 1)][required_columns],
        'Multinucleated_Infected': df[(df['Classify_Mononucleated'] == 0) & (df['Classify_Infected'] == 1)][required_columns],
        'Mononucleated_Uninfected': df[(df['Classify_Mononucleated'] == 1) & (df['Classify_Infected'] == 0)][required_columns],
        'Multinucleated_Uninfected': df[(df['Classify_Mononucleated'] == 0) & (df['Classify_Infected'] == 0)][required_columns]
    }

    # Create the output name
    input_filename = os.path.basename(input_file)
    input_name_no_ext, input_ext = os.path.splitext(input_filename)
    outfile = f"output_{input_name_no_ext}.xlsx"

    # Don't overwrite the output file!
    counter = 1
    original_outfile = outfile
    while os.path.exists(outfile):
        outfile = f"{original_outfile[:-5]}_{counter}.xlsx"
        counter += 1

    # Write categorized data to Excel
    used_sheet_names = set()

    with pd.ExcelWriter(outfile, engine='openpyxl') as writer:
        for sheet_name, data in categories.items():
            # Truncate and make unique
            sheet_name = get_unique_sheet_name(sheet_name)
            data.to_excel(writer, sheet_name=sheet_name, index=False)

    print("")
    print(f"Writing output data to {outfile}. Did you kill her be honest.")
    print("")

    from openpyxl import load_workbook

    # Histogram-style summary of nuclei per infected cell
    infected_nuclei_counts = infected_summary['Children_Nuclei_Count']
    max_bin = args.histogram
    bins = list(range(0, max_bin)) + [float('inf')]
    bin_labels = [str(i) for i in range(1, max_bin)] + [f"{max_bin}+"]
    binned = pd.cut(infected_nuclei_counts,
                    bins=bins,
                    labels=bin_labels,
                    right=True,
                    include_lowest=True)

    nuclei_distribution = binned.value_counts().sort_index().reset_index()
    nuclei_distribution.columns = ['Nuclei per Infected Cell', 'Cell Count']

    # Combine summary and histogram into one sheet with spacing
    combined_sheet = pd.concat([
        infected_summary_df,
        pd.DataFrame([[""] * infected_summary_df.shape[1]] * 2, columns=infected_summary_df.columns),  # empty rows
        pd.DataFrame({"Category": ["Nuclei Distribution in Infected Cells"]}),
        nuclei_distribution.rename(columns={"Nuclei per Infected Cell": "Category", "Cell Count": "Infected Cell Count"})
    ], ignore_index=True)

    # Write to 'Cell Count Summary' sheet
    with pd.ExcelWriter(outfile, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        combined_sheet.to_excel(writer, sheet_name='Cell Count Summary', index=False)

    # Plot a histogram
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, Reference

    # Load workbook and select the sheet
    wb = load_workbook(outfile)
    ws = wb['Cell Count Summary']

    # Find the start of the nuclei distribution table
    for row in ws.iter_rows(min_row=1, max_col=1):
        for cell in row:
            if cell.value == 'Nuclei Distribution in Infected Cells':
                start_row = cell.row + 1  # Data starts on next row
                break

    # Assuming fixed structure (Category in col A, Counts in col B)
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Nuclei per Infected Cell"
    chart.x_axis.title = "Nuclei"
    chart.y_axis.title = "Infected cells"

    # Show axis labels
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"

    # Remove major gridlines from Y axis
    chart.y_axis.majorGridlines = None

    data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + args.histogram - 1)
    cats = Reference(ws, min_col=1, min_row=start_row, max_row=start_row + args.histogram - 1)

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)

    # Position the chart a few rows below the table
    chart_anchor = f"A{start_row + args.histogram + 3}"
    ws.add_chart(chart, chart_anchor)

    wb.save(outfile)

    # Execute the Jane function!
    jane(outfile, args.columns)

    # Reload the workbook to reorder sheets
    wb = load_workbook(outfile)

    # Desired order
    priority_order = ['Cell Count Summary', 'AreaShape_Area Summary', 'Children_Nuclei_Count Sum']

    # Collect existing sheets, preserving order
    existing = wb.sheetnames
    reordered = [s for s in priority_order if s in existing] + [s for s in existing if s not in priority_order]

    # Reorder
    wb._sheets = [wb[s] for s in reordered]

    # Save workbook with new sheet order
    wb.save(outfile)
